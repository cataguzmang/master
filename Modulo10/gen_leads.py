# -*- coding: utf-8 -*-
r"""
Generador de datos sintéticos de leads B2B de productos alimentarios en Latinoamérica.

Ejecutar:
    .\.venv\Scripts\python.exe gen_leads.py

El script pregunta:
  1) Cuántos registros generar (entre 10 y 250).
  2) En qué formato quieres la salida (JSON, CSV o Excel).

Toda la data es 100% sintética. Empresas, personas, correos y teléfonos son ficticios.
"""

import csv
import json
import random
import unicodedata

OUT_DIR = r"C:\Users\caguz\proyectos\claude-work\MASTER\Modulo10"
MIN_REG, MAX_REG = 10, 250

# --- Pools para construir nombres de empresa creíbles -------------------------
PREFIJOS = [
    "Andina de", "Sabores del", "Delicias", "Productos", "Alimentos",
    "Industrias", "Distribuidora", "Comercial", "Grupo", "Conservas",
    "Granos del", "Molinos", "Dulcería", "Frutos", "Bebidas", "Café",
    "Snacks", "NutriVida", "VidaSana", "Abarrotes", "Cosecha",
]
NUCLEOS = [
    "Sur", "Norte", "Pacífico", "Caribe", "Altiplano", "Patagonia", "Andes",
    "Tropicalia", "Sol Naciente", "Cumbre", "Selva Verde", "Mar Azul",
    "La Higuera", "San Jacinto", "Quetzal", "Volcán", "Pampa", "Maya",
    "Polo Austral", "El Trigal", "La Cosecha", "Verde Vida", "Aurora",
    "Sabor Real", "Buen Campo", "Tierra Fértil", "Costa Dorada", "Manantial",
]

# país -> (código telefónico, tlds de dominio típicos)
PAISES = {
    "Colombia": ("+57", [".com", ".co", ".com.co"]),
    "Chile": ("+56", [".cl", ".com"]),
    "México": ("+52", [".mx", ".com.mx", ".com"]),
    "Brasil": ("+55", [".com.br", ".com"]),
    "Perú": ("+51", [".pe", ".com.pe"]),
    "Argentina": ("+54", [".com.ar", ".com"]),
    "Guatemala": ("+502", [".com", ".gt"]),
    "Bolivia": ("+591", [".bo", ".com.bo"]),
    "Ecuador": ("+593", [".ec", ".com.ec"]),
    "Uruguay": ("+598", [".uy", ".com.uy"]),
    "Paraguay": ("+595", [".py", ".com.py"]),
    "Costa Rica": ("+506", [".cr", ".co.cr"]),
    "El Salvador": ("+503", [".sv", ".com.sv"]),
    "Honduras": ("+504", [".hn", ".com"]),
    "Panamá": ("+507", [".com.pa", ".com"]),
    "República Dominicana": ("+1809", [".do", ".com.do"]),
    "Nicaragua": ("+505", [".com.ni", ".com"]),
    "Venezuela": ("+58", [".com.ve", ".com"]),
}

CATEGORIAS = [
    "Snacks", "Bebestible", "No perecible", "Congelados", "Dulces",
    "Salsas", "Café", "Té", "Cereales", "Alimentos saludables",
    "Lácteos", "Panadería", "Conservas",
]

NOMBRES = [
    "Mariana", "Tomás", "Lucía", "Federico", "Camila", "Renzo", "Valeria",
    "Mateo", "Carmen", "Daniela", "Sebastián", "José", "Antonia", "Andrés",
    "Rocío", "Esteban", "Gabriela", "Diego", "Florencia", "Paola", "Rafael",
    "Hugo", "Constanza", "Noelia", "Marcela", "Ignacio", "Valentina", "Joaquín",
    "Isidora", "Martín", "Catalina", "Emiliano", "Sofía", "Nicolás", "Renata",
]
APELLIDOS = [
    "Restrepo", "Fuentealba", "Morales", "Bianchi", "Azevedo", "Cáceres",
    "Hernández", "Quispe", "Vázquez", "Cevallos", "Methol", "Peralta",
    "Vergara", "Mora", "Benítez", "Ocampo", "Xiloj", "Mejía", "Aguirre",
    "Salazar", "Cardoso", "Domínguez", "Pizarro", "Mamani", "Fonseca",
    "Navarro", "Ríos", "Sandoval", "Carvalho", "Paredes", "Villalba",
]


def slug(s):
    """Quita tildes/espacios para emails y dominios."""
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.lower().replace(" ", "")


def preguntar_cantidad():
    while True:
        raw = input(f"¿Cuántos registros quieres generar? (entre {MIN_REG} y {MAX_REG}): ").strip()
        if not raw.isdigit():
            print("  → Ingresa un número válido.")
            continue
        n = int(raw)
        if MIN_REG <= n <= MAX_REG:
            return n
        print(f"  → Número fuera de rango. Elige un valor entre {MIN_REG} y {MAX_REG}.")


def preguntar_formato():
    opciones = {"1": "json", "2": "csv", "3": "excel"}
    while True:
        print("\n¿En qué formato quieres la salida?")
        print("  1) JSON")
        print("  2) CSV")
        print("  3) Excel")
        raw = input("Elige 1, 2 o 3 (o escribe json/csv/excel): ").strip().lower()
        if raw in opciones:
            return opciones[raw]
        if raw in opciones.values():
            return raw
        print("  → Opción no válida. Elige JSON, CSV o Excel.")


def generar(n, seed=42):
    random.seed(seed)
    paises = list(PAISES.keys())
    usados = set()
    leads = []
    while len(leads) < n:
        pais = random.choice(paises)
        dial, tlds = PAISES[pais]
        company = f"{random.choice(PREFIJOS)} {random.choice(NUCLEOS)}"
        if company in usados:
            continue
        usados.add(company)

        first = random.choice(NOMBRES)
        last = random.choice(APELLIDOS)
        domain = slug(company) + random.choice(tlds)
        email = f"{slug(first)}.{slug(last)}@{domain}"
        # Número ficticio: prefijo de abonado en rango no asignable.
        phone = f"{dial} {random.randint(2000, 2999)} {random.randint(1000, 9999)}"

        leads.append({
            "company_name": company,
            "country": pais,
            "category": random.choice(CATEGORIAS),
            "contact_name": f"{first} {last}",
            "contact_email": email,
            "contact_phone": phone,
            "exports": random.choice([True, False]),
        })
    return leads


def guardar_json(leads):
    out = OUT_DIR + r"\synthetic_leads.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(leads, f, ensure_ascii=False, indent=2)
    return out


def guardar_csv(leads):
    out = OUT_DIR + r"\synthetic_leads.csv"
    campos = ["company_name", "country", "category", "contact_name",
              "contact_email", "contact_phone", "exports"]
    with open(out, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=campos)
        w.writeheader()
        for row in leads:
            r = dict(row)
            r["exports"] = "true" if r["exports"] else "false"
            w.writerow(r)
    return out


def guardar_excel(leads):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    out = OUT_DIR + r"\synthetic_leads.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "synthetic_leads"
    headers = ["company_name", "country", "category", "contact_name",
               "contact_email", "contact_phone", "exports"]
    ws.append(headers)
    for row in leads:
        ws.append([row[h] for h in headers])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5496")
    for col, w in zip("ABCDEFG", [26, 22, 22, 22, 40, 20, 10]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(out)
    return out


def main():
    print("=== Generador de leads sintéticos B2B (alimentos · Latam) ===\n")
    n = preguntar_cantidad()
    fmt = preguntar_formato()
    leads = generar(n)

    if fmt == "json":
        out = guardar_json(leads)
    elif fmt == "csv":
        out = guardar_csv(leads)
    else:
        out = guardar_excel(leads)

    print(f"\n✓ Generados {len(leads)} registros en formato {fmt.upper()}.")
    print(f"✓ Guardado en: {out}")


if __name__ == "__main__":
    main()
